#Noah Kruse

import openpyxl

from openpyxl import workbook, load_workbook

wb = load_workbook('Stocks.xlsx')
ws = wb.active

wb1 = load_workbook("Industries.xlsx")
ws1 = wb1.active

class Stock:
    def __init__(self, row):
        self.Name = ws['C'+str(row)].value
        self.Industry = ws['E'+str(row)].value
        self.PE = ws['F'+str(row)].value
        self.FPE = ws['G'+str(row)].value
        self.PS = ws['N'+str(row)].value
        self.FPS = ws['O'+str(row)].value
        self.PB = ws['P'+str(row)].value
        self.PEG = ws['R'+str(row)].value
        self.Rating = ws['J'+str(row)].value

    def __str__(self):
        return self.Name+ "\n"+ "\n"+ "PE Ratio   : " + str(self.PE)+ "\n"+"FPE Ratio  : " + str(self.FPE)+ "\n"+"PS Ratio   : " + str(self.PS)+ "\n"+"FPS Ratio  : " + str(self.FPS)+ "\n"+"PB Ratio   : " + str(self.PB)+ "\n"+"PEG Ratio  : " + str(self.PEG)+ "\n"+"Rating     : " + str(self.Rating)+ "\n"

class Industry:
    def __init__(self, name):
        self.Name = name
        self.PE = ws1['D'+str(Industry.findRow(name))].value
        self.PS = Industry.findPS(name)

    def findRow(name):
        for i in range(2, 147):
            if name in ws1['A'+str(i)].value:
                return i
    def findPS(name):
        numCompanies = 0
        totalPS = 0
        for i in range(2, 5807):
            try:
                if name in ws['E'+str(i)].value:
                    numCompanies = numCompanies + 1
                    totalPS = totalPS + ws['N'+str(i)].value
            except:
                pass
        return totalPS/numCompanies
    
    def __str__(self):
        return "Industry: "+self.Name+ "\n"+ "\n"+ "PE Ratio   : " + str(self.PE)+ "\n"+"PS Ratio   : " + str(self.PS)+ "\n"
                
def SearchStock():

    StockName = (input("Enter Stock Name: ")).lower()
    print('\n')
    Possible = []
    row = -1

    for i in range(2, 5807):
        curr = ws['C'+ str(i)].value
        
        if StockName in curr.lower():
            Possible.append([curr, i])
    
    if len(Possible) == 0: 
        print('No Stock Found')
        print('\n')
    else:   
        print("Stocks found:")
        for i in range(len(Possible)):
            print(str(i +1) + ": "+ Possible[i][0])
        print('\n')
        response = input("Enter Number: ")
        row = Possible[int(response) - 1][1]
        print('\n')
        stock = Stock(row)
        print(stock)
        print(Industry(stock.Industry))
    
    if row != -1:
        if input("Would you like a recommendation?(y/n): ") == "y":
            FullEval(stock)
            

    if input("Analyze another Stock?(y/n): ") == "y":
        SearchStock()
        
def FullEval(stock):
    score = 0
    test = 0
    ind = Industry(stock.Industry)
    try:
        score = PEEval(stock, ind, score)
        test = test +1
    except:
        pass
    try:
        score = FPEEval(stock, ind, score)
        test = test +1
    except:
        pass
    try:
        score = PSEval(stock, ind, score)
        test = test +1
    except:
        pass
    try:
        score = FPSEval(stock, ind, score)
        test = test +1
    except:
        pass
    try:
        score = PBEval(stock, ind, score)
        test = test +1
    except:
        pass
    try:
        score = PEGEval(stock, ind, score)
        test = test +1
    except:
        pass
    score = round(score *5 /test, 1)
    score = score +5
    if score > 5:
        print("Tests ran:" +str(test))
        print("This stock is a good investment.\n\nRating: "+str(score)+"/10" )
    elif score < 5:
        print("Tests ran:" +str(test))
        print("This stock is a bad investment.\n\nRating: "+str(score)+"/10" )
    else:
        print("Tests ran:" +str(test))
        print("This stock is mid.\n\nRating: "+str(score)+"/10" )

def PEEval(stock, ind, score):
    if ind.PE > stock.PE:
        score = score + 1
    elif ind.PE < stock.PE:
        score = score - 1
    return score
        
def FPEEval(stock, ind, score):
    if ind.PE > stock.FPE:
        score = score + 1
    elif ind.PE < stock.FPE:
        score = score - 1
    return score
        
def PSEval(stock, ind, score):
    if ind.PS > stock.PS:
        score = score + 1
    elif ind.PS < stock.PS:
        score = score - 1
    return score
        
def FPSEval(stock, ind, score):
    if ind.PS > stock.FPS:
        score = score + 1
    elif ind.PS < stock.FPS:
        score = score - 1
    return score

def PBEval(stock, score):
    if stock.PB < 1:
        score = score + 1
    elif stock.PB > 1:
        score = score - 1
    return score

def PEGEval(stock, score):
    if stock.PEG < 1:
        score = score + 1
    elif stock.PEG > 1:
        score = score - 1
    return score

SearchStock()

